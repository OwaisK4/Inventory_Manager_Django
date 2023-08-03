from typing import Any, Dict
from django.forms.models import BaseModelForm
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic import ListView, DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView
from django.urls import reverse, reverse_lazy
from datetime import datetime
from django.utils import timezone
from .models import Asset, Employee, Category, Manufacturer, Department, Status, Attachement, Supplier, Maintenance, Accessory, Location, Checkout, Activity, Audit, ScheduledAudit, Change, License
from .forms import AssetModelForm, EmployeeModelForm, CategoryModelForm, ManufacturerModelForm, AttachementModelForm, SupplierModelForm, DepartmentModelForm, StatusModelForm, MaintenanceModelForm, AccessoryModelForm, LocationModelForm, RegistrationForm, CheckoutModelForm, AuditModelForm, ScheduledAuditModelForm, LicenseModelForm
# from django.contrib.auth import login, authenticate
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _
from django.views.decorators.csrf import csrf_exempt
from django.db.models import ForeignKey
import qrcode
from io import BytesIO
from base64 import b64encode
import openpyxl

# Create your views here.
def index(request):
    return HttpResponse("<H2>HEY! Welcome to my website! </H2>")

@login_required
def master_page(request):
    return render(request, 'master.html')

@login_required
def dashboard(request):
    asset_count = Asset.objects.all().count()
    accessories_count = Accessory.objects.all().count()
    user_count = User.objects.all().count()
    audits_count = ScheduledAudit.objects.all().count()
    checked_in_count = Asset.objects.filter(checkout_status='I').count()
    checked_out_count = Asset.objects.filter(checkout_status='O').count()
    activities = Activity.objects.all().order_by('-id')[:10]

    total_asset_cost = 0
    for asset in Asset.objects.all():
        total_asset_cost += asset.price
    total_accessory_cost = 0
    for accessory in Accessory.objects.all():
        total_accessory_cost += accessory.price

    context = {
        'asset_count': asset_count,
        'accessories_count': accessories_count,
        'total_asset_cost': total_asset_cost,
        'total_accessory_cost': total_accessory_cost,
        'user_count': user_count,
        'audits_count': audits_count,
        'checked_in_count': checked_in_count,
        'checked_out_count': checked_out_count,
        'activities': activities,
    }
    return render(request, 'webapp/dashboard.html', context)

@login_required
def about(request):
    return render(request, 'registration/about.html')

@csrf_exempt
def webhook(request):
    if request.method == 'POST':
        print(request.body)
        data = request.POST
        # pk = data.get("ID")
        # if Asset.objects.filter(pk=pk).exists():
        #     asset = Asset.objects.get(pk=pk)
        #     change = Change()
        #     change.asset = asset
        #     change.status = data.get("Status")
        #     change.user = data.get("User")
        #     change.processor = data.get("Processor")
        #     change.RAM = data.get("RAM")
        #     change.disk_space = data.get("sum_total_disk")
        #     change.change_in_processor = data.get("status")
        #     change.change_in_ram = data.get("status")
        #     change.change_in_disk = data.get("status")
        #     change.message = data.get("status")
        return HttpResponse("Webhook received!")

@login_required
def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            # username = form.cleaned_data.get('username')
            # raw_password = form.cleaned_data.get('password1')
            # user = authenticate(username=username, password=raw_password)
            # login(request, user)
            return HttpResponseRedirect(reverse('dashboard'))
    else:
        form = RegistrationForm()
    return render(request, 'registration/register.html', {'form': form})

class UserListView(LoginRequiredMixin, ListView):
    model = User
    context_object_name = 'users'
    template_name = 'registration/view_users.html'

@login_required
def input_inventory_form(request):
    if request.method == 'POST':
        form = AssetModelForm(request.POST, request.FILES)
        if form.is_valid():
            if form.cleaned_data.get("price") < 0:
                form.price = 0
                # raise ValidationError(_("Price cannot be less than 0."))
            created_asset = form.save(commit=False)
            activity = Activity(event="Create", user=request.user)
            activity.asset = created_asset
            activity.asset_string = str(created_asset)
            activity.employee = created_asset.assigned_to
            created_asset.save()
            activity.save()
            redirect_to = request.POST.get('next', '')
            if redirect_to:
                return HttpResponseRedirect(redirect_to)
            else:
                return HttpResponseRedirect(reverse('view_inventory-list'))
    else:
        form = AssetModelForm(request.POST)
    context = {
        'form': form,
    }
    return render(request, 'webapp/input_inventory.html', context)

@login_required
def input_employee_form(request):
    if request.method == 'POST':
        form = EmployeeModelForm(request.POST)
        if form.is_valid():
            form.save(commit=True)
            redirect_to = request.POST.get('next', '')
            if redirect_to:
                return HttpResponseRedirect(redirect_to)
            else:
                return HttpResponseRedirect(reverse('view_employee-list'))
    else:
        form = EmployeeModelForm(request.POST)
    context = {
        'form': form,
        'next': request.GET.get('next', '')
    }
    return render(request, 'webapp/input_employee.html', context)

class InventoryListView(LoginRequiredMixin, ListView):
    model = Asset
    context_object_name = 'inventory'
    template_name = 'webapp/view_inventory.html'

class InventoryDetailView(LoginRequiredMixin, DetailView):
    model = Asset
    context_object_name = 'asset'
    template_name = 'webapp/view_inventory_detail.html'

class InventoryDeleteView(LoginRequiredMixin, DeleteView):
    model = Asset
    fields = '__all__'
    context_object_name = 'asset'
    template_name = 'webapp/view_inventory_confirm_delete.html'
    success_url = reverse_lazy('view_inventory-list')

    def form_valid(self, form) -> HttpResponse:
        created_asset = self.get_object()
        activity = Activity(event="Delete", user=self.request.user)
        activity.asset = created_asset
        activity.asset_string = str(created_asset)
        activity.employee = created_asset.assigned_to
        activity.save()
        return super().form_valid(form)

class InventoryUpdateView(LoginRequiredMixin, UpdateView):
    model = Asset
    context_object_name = 'asset'
    form_class = AssetModelForm
    template_name = 'webapp/input_inventory.html'
    success_url = reverse_lazy('view_inventory-list')
    
    def form_valid(self, form) -> HttpResponse:
        created_asset = form.save(commit=False)
        activity = Activity(event="Update", user=self.request.user)
        activity.asset = created_asset
        activity.asset_string = str(created_asset)
        activity.employee = created_asset.assigned_to
        changes = form.changed_data
        notes = ""
        if changes:
            notes = "Changed "
            for change in changes:
                # initial = form.get_initial_for_field(form.fields[change], change)
                initial = form[change].initial
                field_object = Asset._meta.get_field(change)
                current = form.cleaned_data[change]
                # print(f"Initial = {initial}")
                # print(f"Current = {current}")
                notes += change
                if not isinstance(field_object, ForeignKey):
                    notes += " from "
                    notes += str(initial)
                notes += " to "
                notes += str(current)
                notes += ", "
            notes = notes[:-2]
        activity.notes = notes
        # print(form.get_initial_for_field(form.fields["price"], "price"))
        # activity.notes = changes
        activity.save()
        return super().form_valid(form)

@login_required
def generate_barcode_asset(request, pk):
    asset = Asset.objects.get(pk=pk)
    barcode = qrcode.make("LDPAST " + str(asset.id) + " " + asset.name)
    response = HttpResponse(content_type='image/jpg')
    barcode.save(response, "JPEG")
    # response['Content-Disposition'] = 'attachment; filename="piece.jpg"'
    return response

@login_required
def generate_all_barcodes_asset(request):
    assets = Asset.objects.all()
    images = []
    for asset in assets:
        barcode = qrcode.make("LDPAST " + str(asset.id) + " " + asset.name)
        image = BytesIO()
        barcode.save(image, format="JPEG")
        dataurl = 'data:image/png;base64,' + b64encode(image.getvalue()).decode('ascii')
        images.append(dataurl)

    context = {
        'images': images
    }
    return render(request, 'webapp/view_barcodes.html', context)

@login_required
def generate_barcode_accessory(request, pk):
    accessory = Accessory.objects.get(pk=pk)
    barcode = qrcode.make("LDPASC " + str(accessory.id) + " " + str(accessory))
    response = HttpResponse(content_type='image/jpg')
    barcode.save(response, "JPEG")
    # response['Content-Disposition'] = 'attachment; filename="piece.jpg"'
    return response

@login_required
def generate_all_barcodes_accessory(request):
    accessories = Accessory.objects.all()
    images = []
    for accessory in accessories:
        barcode = qrcode.make("LDPASC " + str(accessory.id) + " " + str(accessory))
        image = BytesIO()
        barcode.save(image, format="JPEG")
        dataurl = 'data:image/png;base64,' + b64encode(image.getvalue()).decode('ascii')
        images.append(dataurl)
        
    context = {
        'images': images
    }
    return render(request, 'webapp/view_barcodes.html', context)

# @login_required
# def inventory_checkout(request, pk):
#     asset = Asset.objects.get(pk=pk)
#     if asset.checkout_status == 'I':
#         Asset.objects.filter(pk=pk).update(checkout_status='O')
#     else:
#         Asset.objects.filter(pk=pk).update(checkout_status='I')
#     return HttpResponseRedirect(reverse('view_inventory-list'))

@login_required
def inventory_attachements_list(request, pk):
    asset = Asset.objects.get(pk=pk)
    attachements = Attachement.objects.filter(asset=asset)
    context = {
        'asset': asset,
        'attachements': attachements,
    }
    return render(request, 'webapp/view_inventory_attachements.html', context)

@login_required
def inventory_attachements_delete(request, pk, pk_attachement):
    attachement = Attachement.objects.get(pk=pk_attachement)
    attachement.delete()
    return HttpResponseRedirect(reverse('view_inventory-attachements-list', args=(str(pk),)))

class AttachementDeleteView(LoginRequiredMixin, DeleteView):
    model = Attachement
    fields = '__all__'
    context_object_name = 'attachement'
    template_name = 'webapp/generic_confirm_delete.html'
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "Attachement"
        context["delete_object"] = f"attachement # {self.get_object().id}"
        return context
    def get_success_url(self) -> str:
        pk_asset = self.kwargs.get("pk_asset", "")
        if pk_asset:
            return reverse("view_inventory-attachements-list", args=(str(pk_asset),))
        else:
            return reverse("view_inventory-list")

class InventoryAttachementsAddView(LoginRequiredMixin, CreateView):
    model = Attachement
    form_class = AttachementModelForm
    template_name = 'webapp/view_inventory_input_attachements.html'

    def get_success_url(self) -> str:
        if self.kwargs['pk']:
            return reverse('view_inventory-attachements-list', args=str(self.kwargs['pk'],))
        else:
            return super().get_success_url()

    def form_valid(self, form):
        attachement = form.save(commit=False)
        attachement.asset = Asset.objects.get(pk=self.kwargs['pk'])
        attachement.save()
        return HttpResponseRedirect(self.get_success_url())

@login_required
def inventory_maintenances_list(request, pk):
    asset = Asset.objects.get(pk=pk)
    maintenances = Maintenance.objects.filter(asset=asset)
    context = {
        'asset': asset,
        'maintenances': maintenances,
    }
    return render(request, 'webapp/view_inventory_maintenances.html', context)

@login_required
def inventory_maintenances_delete(request, pk, pk_maintenance):
    Maintenance.objects.filter(pk=pk_maintenance).delete()
    return HttpResponseRedirect(reverse('view_inventory-maintenances-list', args=(str(pk),)))

class MaintenanceDeleteView(LoginRequiredMixin, DeleteView):
    model = Maintenance
    fields = '__all__'
    context_object_name = 'maintenance'
    template_name = 'webapp/generic_confirm_delete.html'
    # success_url = reverse_lazy('view_supplier-list')
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "Maintenance"
        context["delete_object"] = f"maintenance: {self.get_object()}"
        return context
    def get_success_url(self) -> str:
        pk_asset = self.kwargs.get("pk_asset", "")
        if pk_asset:
            return reverse("view_inventory-maintenances-list", args=(str(pk_asset),))
        else:
            return reverse("view_inventory-list")

class InventoryMaintenancesAddView(LoginRequiredMixin, CreateView):
    model = Maintenance
    form_class = MaintenanceModelForm
    template_name = 'webapp/view_inventory_input_maintenances.html'

    def get_success_url(self) -> str:
        if self.kwargs['pk']:
            return reverse('view_inventory-maintenances-list', args=(self.kwargs['pk'],))
        else:
            return super().get_success_url()

    def form_valid(self, form):
        maintenance = form.save(commit=False)
        maintenance.asset = Asset.objects.get(pk=self.kwargs['pk'])
        maintenance.save()
        return HttpResponseRedirect(self.get_success_url())

@login_required
def inventory_checkout_list(request, pk):
    asset = Asset.objects.get(pk=pk)
    checkouts = Checkout.objects.filter(asset=asset)
    context = {
        'asset': asset,
        'checkouts': checkouts,
    }
    return render(request, 'webapp/view_inventory_checkout_log.html', context)

class CheckoutView(LoginRequiredMixin, CreateView):
    model = Checkout
    form_class = CheckoutModelForm
    template_name = 'webapp/view_inventory_checkout.html'

    def get_success_url(self) -> str:
        next_url = self.request.GET.get('next', None)
        if next_url:
            return str(next_url)
        else:
            return reverse_lazy('view_inventory-checkout-list', args=str(self.kwargs['pk'],))

    def form_valid(self, form):
        checkout = form.save(commit=False)
        checkout.asset = Asset.objects.get(pk=self.kwargs['pk'])
        checkout.type = 'O'
        checkout.user = self.request.user
        checkout.save()

        activity = Activity(event="Check-out", user=self.request.user)
        activity.asset = checkout.asset
        activity.asset_string = str(checkout.asset)
        activity.employee = checkout.asset.assigned_to
        activity.save()
        
        Asset.objects.filter(pk=self.kwargs['pk']).update(checkout_status='O')
        return HttpResponseRedirect(self.get_success_url())

class CheckinView(LoginRequiredMixin, CreateView):
    model = Checkout
    form_class = CheckoutModelForm
    template_name = 'webapp/view_inventory_checkin.html'
    
    def get_success_url(self) -> str:
        next_url = self.request.GET.get('next', None)
        if next_url:
            return str(next_url)
        else:
            return reverse_lazy('view_inventory-checkout-list', args=str(self.kwargs['pk'],))
        
    def form_valid(self, form):
        checkin = form.save(commit=False)
        checkin.asset = Asset.objects.get(pk=self.kwargs['pk'])
        checkin.type = 'I'
        checkin.user = self.request.user
        checkin.save()

        activity = Activity(event="Check-in", user=self.request.user)
        activity.asset = checkin.asset
        activity.asset_string = str(checkin.asset)
        activity.employee = checkin.asset.assigned_to
        activity.save()

        Asset.objects.filter(pk=self.kwargs['pk']).update(checkout_status='I')
        return HttpResponseRedirect(self.get_success_url())
    
@login_required
def inventory_audit_list(request, pk):
    asset = Asset.objects.get(pk=pk)
    audits = Audit.objects.filter(asset=asset)
    context = {
        'asset': asset,
        'audits': audits,
    }
    return render(request, 'webapp/view_inventory_audit_log.html', context)

class ScheduledAuditDeleteView(LoginRequiredMixin, DeleteView):
    model = ScheduledAudit
    fields = '__all__'
    context_object_name = 'audit'
    template_name = 'webapp/generic_confirm_delete.html'
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "audit"
        context["delete_object"] = str(self.get_object())
        return context
    def get_success_url(self) -> str:
        return reverse('view_inventory-scheduled-audit-list')

class InventoryAuditView(LoginRequiredMixin, CreateView):
    model = Audit
    form_class = AuditModelForm
    template_name = 'webapp/view_inventory_audit.html'

    def get_success_url(self) -> str:
        if self.kwargs['pk']:
            return reverse('view_inventory-audit-schedule', args=(self.kwargs['pk'],))
        else:
            return super().get_success_url()

    def form_valid(self, form):
        audit = form.save(commit=False)
        asset = Asset.objects.get(pk=self.kwargs['pk'])
        audit.asset = asset
        audit.asset_string = str(asset)
        audit.user = self.request.user
        audit.save()
        return HttpResponseRedirect(self.get_success_url())
    
class InventoryAuditScheduleView(LoginRequiredMixin, CreateView):
    model = ScheduledAudit
    form_class = ScheduledAuditModelForm
    template_name = 'webapp/view_inventory_audit_schedule.html'

    def get_success_url(self) -> str:
        if self.kwargs['pk']:
            return reverse('view_inventory-audit-list', args=(self.kwargs['pk'],))
        else:
            return super().get_success_url()

    def form_valid(self, form):
        scheduled_audit = form.save(commit=False)
        asset = Asset.objects.get(pk=self.kwargs['pk'])
        if asset.scheduled_audit.exists():
            asset.scheduled_audit.all().delete()
        scheduled_audit.asset = asset
        scheduled_audit.save()
        return HttpResponseRedirect(self.get_success_url())

class ScheduledAuditsListView(LoginRequiredMixin, ListView):
    model = ScheduledAudit
    context_object_name = 'scheduled_audits'
    template_name = 'webapp/view_inventory_scheduled_audits.html'

@login_required
def scheduled_audit_view(request, pk):
    audit = ScheduledAudit.objects.get(pk=pk)
    pk_asset = audit.asset.id
    # audit.delete()
    return HttpResponseRedirect(reverse('view_inventory-audit', args=(pk_asset,)))

class ActivityListView(LoginRequiredMixin, ListView):
    model = Activity
    context_object_name = 'activities'
    template_name = 'webapp/view_activity_log.html'

class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    context_object_name = 'employees'
    template_name = 'webapp/view_employees.html'

# class EmployeeDetailView(DetailView):
#     model = Employee
#     context_object_name = 'employee'
#     template_name = 'webapp/view_employee_detail.html'

class EmployeeDeleteView(LoginRequiredMixin, DeleteView):
    model = Employee
    fields = '__all__'
    context_object_name = 'employee'
    template_name = 'webapp/view_employee_confirm_delete.html'
    success_url = reverse_lazy('view_employee-list')

class EmployeeUpdateView(LoginRequiredMixin, UpdateView):
    model = Employee
    context_object_name = 'employee'
    form_class = EmployeeModelForm
    template_name = 'webapp/input_employee.html'
    success_url = reverse_lazy('view_employee-list')

@login_required
def time_view(request):
    now = datetime.now()
    time = f"Current time is {now}"
    return HttpResponse(time)

# Views for Category objects
class CategoryListView(LoginRequiredMixin, ListView):
    model = Category
    fields = '__all__'
    context_object_name = 'categories'
    template_name = 'webapp/view_categories.html'

class CategoryCreateView(LoginRequiredMixin, CreateView):
    model = Category
    form_class = CategoryModelForm
    template_name = 'webapp/input_category.html'
    def get_success_url(self) -> str:
        next_url = self.request.GET.get('next', None)
        if next_url:
            return str(next_url)
        else:
            return reverse_lazy('view_category-list')

@login_required
def category_delete(request, pk):
    Category.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(reverse('view_category-list'))

class CategoryDeleteView(LoginRequiredMixin, DeleteView):
    model = Category
    fields = '__all__'
    context_object_name = 'category'
    template_name = 'webapp/generic_confirm_delete.html'
    success_url = reverse_lazy('view_category-list')
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "Category"
        context["delete_object"] = f"category: {self.get_object().name}"
        return context

class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    model = Category
    form_class = CategoryModelForm
    template_name = 'webapp/input_category.html'
    success_url = reverse_lazy('view_category-list')

# Views for Manufacturer objects
class ManufacturerListView(LoginRequiredMixin, ListView):
    model = Manufacturer
    fields = '__all__'
    context_object_name = 'manufacturers'
    template_name = 'webapp/view_manufacturers.html'

class ManufacturerCreateView(LoginRequiredMixin, CreateView):
    model = Manufacturer
    form_class = ManufacturerModelForm
    template_name = 'webapp/input_manufacturer.html'
    def get_success_url(self) -> str:
        next_url = self.request.GET.get('next', None)
        if next_url:
            return str(next_url)
        else:
            return reverse_lazy('view_manufacturer-list')

@login_required
def manufacturer_delete(request, pk):
    Manufacturer.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(reverse('view_manufacturer-list'))

class ManufacturerDeleteView(LoginRequiredMixin, DeleteView):
    model = Manufacturer
    fields = '__all__'
    context_object_name = 'manufacturer'
    template_name = 'webapp/generic_confirm_delete.html'
    success_url = reverse_lazy('view_manufacturer-list')
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "Manufacturer"
        context["delete_object"] = f"manufacturer: {self.get_object().name}"
        return context

class ManufacturerUpdateView(LoginRequiredMixin, UpdateView):
    model = Manufacturer
    form_class = ManufacturerModelForm
    template_name = 'webapp/input_manufacturer.html'
    success_url = reverse_lazy('view_manufacturer-list')

# Views for Supplier objects
class SupplierListView(LoginRequiredMixin, ListView):
    model = Supplier
    fields = '__all__'
    context_object_name = 'suppliers'
    template_name = 'webapp/view_suppliers.html'

class SupplierCreateView(LoginRequiredMixin, CreateView):
    model = Supplier
    form_class = SupplierModelForm
    template_name = 'webapp/input_supplier.html'
    def get_success_url(self) -> str:
        next_url = self.request.GET.get('next', None)
        if next_url:
            return str(next_url)
        else:
            return reverse_lazy('view_supplier-list')

@login_required
def supplier_delete(request, pk):
    Supplier.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(reverse('view_supplier-list'))

class SupplierDeleteView(LoginRequiredMixin, DeleteView):
    model = Supplier
    fields = '__all__'
    context_object_name = 'supplier'
    template_name = 'webapp/generic_confirm_delete.html'
    success_url = reverse_lazy('view_supplier-list')
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "Supplier"
        context["delete_object"] = f"supplier: {self.get_object().name}"
        return context

class SupplierUpdateView(LoginRequiredMixin, UpdateView):
    model = Supplier
    form_class = SupplierModelForm
    template_name = 'webapp/input_supplier.html'
    success_url = reverse_lazy('view_supplier-list')

# Views for Department objects
class DepartmentListView(LoginRequiredMixin, ListView):
    model = Department
    fields = '__all__'
    context_object_name = 'departments'
    template_name = 'webapp/view_departments.html'

class DepartmentCreateView(LoginRequiredMixin, CreateView):
    model = Department
    form_class = DepartmentModelForm
    template_name = 'webapp/input_department.html'
    def get_success_url(self) -> str:
        next_url = self.request.GET.get('next', None)
        if next_url:
            return str(next_url)
        else:
            return reverse_lazy('view_department-list')

@login_required
def department_delete(request, pk):
    Department.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(reverse('view_department-list'))

class DepartmentDeleteView(LoginRequiredMixin, DeleteView):
    model = Department
    fields = '__all__'
    context_object_name = 'department'
    template_name = 'webapp/generic_confirm_delete.html'
    success_url = reverse_lazy('view_department-list')
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "Department"
        context["delete_object"] = f"department: {self.get_object().name}"
        return context

class DepartmentUpdateView(LoginRequiredMixin, UpdateView):
    model = Department
    form_class = DepartmentModelForm
    template_name = 'webapp/input_department.html'
    success_url = reverse_lazy('view_department-list')

# Views for Status objects
class StatusListView(LoginRequiredMixin, ListView):
    model = Status
    fields = '__all__'
    context_object_name = 'statuses'
    template_name = 'webapp/view_statuses.html'

class StatusCreateView(LoginRequiredMixin, CreateView):
    model = Status
    form_class = StatusModelForm
    template_name = 'webapp/input_status.html'
    def get_success_url(self) -> str:
        next_url = self.request.GET.get('next', None)
        if next_url:
            return str(next_url)
        else:
            return reverse_lazy('view_status-list')

@login_required
def status_delete(request, pk):
    Status.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(reverse('view_status-list'))

class StatusDeleteView(LoginRequiredMixin, DeleteView):
    model = Status
    fields = '__all__'
    context_object_name = 'status'
    template_name = 'webapp/generic_confirm_delete.html'
    success_url = reverse_lazy('view_status-list')
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "Status"
        context["delete_object"] = f"status: {self.get_object().name}"
        return context

class StatusUpdateView(LoginRequiredMixin, UpdateView):
    model = Status
    form_class = StatusModelForm
    template_name = 'webapp/input_status.html'
    success_url = reverse_lazy('view_status-list')

# Views for Accessory objects
class AccessoryCreateView(LoginRequiredMixin, CreateView):
    model = Accessory
    form_class = AccessoryModelForm
    template_name = 'webapp/input_accessory.html'
    def get_success_url(self) -> str:
        next_url = self.request.GET.get('next', None)
        if next_url:
            return str(next_url)
        else:
            return reverse_lazy('view_accessory-list')

class AccessoryListView(LoginRequiredMixin, ListView):
    model = Accessory
    context_object_name = 'accessories'
    template_name = 'webapp/view_accessory.html'

class AccessoryDetailView(LoginRequiredMixin, DetailView):
    model = Accessory
    context_object_name = 'accessory'
    template_name = 'webapp/view_accessory_detail.html'

class AccessoryDeleteView(LoginRequiredMixin, DeleteView):
    model = Accessory
    fields = '__all__'
    context_object_name = 'accessory'
    template_name = 'webapp/view_accessory_confirm_delete.html'
    success_url = reverse_lazy('view_accessory-list')

class AccessoryUpdateView(LoginRequiredMixin, UpdateView):
    model = Accessory
    context_object_name = 'accessory'
    form_class = AccessoryModelForm
    template_name = 'webapp/input_accessory.html'
    success_url = reverse_lazy('view_accessory-list')

# Views for License objects
class LicenseCreateView(LoginRequiredMixin, CreateView):
    model = License
    form_class = LicenseModelForm
    template_name = 'webapp/input_license.html'
    success_url = reverse_lazy('view_license-list')

class LicenseListView(LoginRequiredMixin, ListView):
    model = License
    context_object_name = 'licenses'
    template_name = 'webapp/view_license.html'

class LicenseDetailView(LoginRequiredMixin, DetailView):
    model = License
    context_object_name = 'license'
    template_name = 'webapp/view_license_detail.html'

class LicenseDeleteView(LoginRequiredMixin, DeleteView):
    model = License
    fields = '__all__'
    context_object_name = 'license'
    template_name = 'webapp/generic_confirm_delete.html'
    success_url = reverse_lazy('view_license-list')
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["delete_type"] = "License"
        context["delete_object"] = f"license: {self.get_object().software_name}"
        return context

class LicenseUpdateView(LoginRequiredMixin, UpdateView):
    model = License
    context_object_name = 'license'
    form_class = LicenseModelForm
    template_name = 'webapp/input_license.html'
    success_url = reverse_lazy('view_license-list')

# Views for Location objects
class LocationCreateandDisplayView(LoginRequiredMixin, CreateView):
    model = Location
    form_class = LocationModelForm
    template_name = 'webapp/input_and_display_location.html'
    success_url = reverse_lazy('input_and_display_location')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        locations = Location.objects.all()
        context['locations'] = locations
        return context

class LocationListView(LoginRequiredMixin, ListView):
    model = Location
    context_object_name = 'locations'
    template_name = 'webapp/view_location.html'

class LocationDetailView(LoginRequiredMixin, DetailView):
    model = Location
    context_object_name = 'location'
    template_name = 'webapp/view_location_detail.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['assets'] = Asset.objects.filter(location=self.get_object())
        return context

class LocationDeleteView(LoginRequiredMixin, DeleteView):
    model = Location
    fields = '__all__'
    context_object_name = 'location'
    template_name = 'webapp/view_location_confirm_delete.html'
    success_url = reverse_lazy('input_and_display_location')

# class LocationUpdateView(UpdateView):
#     model = Location
#     context_object_name = 'location'
#     form_class = LocationModelForm
#     template_name = 'webapp/input_location.html'
#     success_url = reverse_lazy('view_location-list')

def export_assets(request):
    assets = Asset.objects.all()
    # fields = Asset._meta.get_fields()
    all_fields = [f.name for f in Asset._meta.get_fields()]

    # return HttpResponse(" ".join(all_fields))
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet["A1"] = "Category"
    sheet["B1"] = "Manufacturer"
    sheet["C1"] = "Name"
    sheet["D1"] = "Price"
    sheet["E1"] = "Department"
    sheet["F1"] = "Assigned to"
    sheet["G1"] = "Location"
    sheet["H1"] = "Purchase_date"
    sheet["I1"] = "Processor"
    sheet["J1"] = "RAM"
    sheet["K1"] = "HDD"
    sheet["L1"] = "SSD"
    sheet["M1"] = "Checkout status"
    sheet["N1"] = "Status"
    sheet["O1"] = "supplier"
    sheet["P1"] = "Serial"
    sheet["Q1"] = "Invoice"
    
    for i in range(len(assets)):
        if assets[i].category:
            sheet.cell(row = i+2, column = 1).value = assets[i].category.name
        if assets[i].manufacturer:
            sheet.cell(row = i+2, column = 2).value = assets[i].manufacturer.name
        sheet.cell(row = i+2, column = 3).value = assets[i].name
        sheet.cell(row = i+2, column = 4).value = assets[i].price
        if assets[i].department:
            sheet.cell(row = i+2, column = 5).value = assets[i].department.name
        if assets[i].assigned_to:
            sheet.cell(row = i+2, column = 6).value = assets[i].assigned_to.name
        if assets[i].location:
            sheet.cell(row = i+2, column = 7).value = assets[i].location.name
        sheet.cell(row = i+2, column = 8).value = assets[i].purchase_date
        sheet.cell(row = i+2, column = 9).value = assets[i].processor
        sheet.cell(row = i+2, column = 10).value = assets[i].ram
        sheet.cell(row = i+2, column = 11).value = assets[i].hdd
        sheet.cell(row = i+2, column = 12).value = assets[i].ssd
        sheet.cell(row = i+2, column = 13).value = assets[i].get_checkout_status_display()
        if assets[i].status:
            sheet.cell(row = i+2, column = 14).value = assets[i].status.name
        if assets[i].supplier:
            sheet.cell(row = i+2, column = 15).value = assets[i].supplier.name
        sheet.cell(row = i+2, column = 16).value = assets[i].serial
        sheet.cell(row = i+2, column = 17).value = assets[i].invoice

    workbook.save(object)

    # create a response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    # tell the browser what the file is named
    response['Content-Disposition'] = 'attachment;filename="Exported_assets.xlsx"'
    # put the spreadsheet data into the response
    response.write(object.getvalue())
    # return the response
    return response

def export_accessories(request):
    accessories = Accessory.objects.all()
    all_fields = [f.name for f in Asset._meta.get_fields()]

    # return HttpResponse(" ".join(all_fields))
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet["A1"] = "Category"
    sheet["B1"] = "Manufacturer"
    sheet["C1"] = "Name"
    sheet["D1"] = "Price"
    sheet["E1"] = "Department"
    sheet["F1"] = "Assigned to"
    sheet["G1"] = "Location"
    sheet["H1"] = "Purchase date"
    sheet["I1"] = "Status"
    sheet["J1"] = "Supplier"
    sheet["K1"] = "Serial"
    sheet["L1"] = "Invoice"
    
    for i in range(len(accessories)):
        if accessories[i].category:
            sheet.cell(row = i+2, column = 1).value = accessories[i].category.name
        if accessories[i].manufacturer:
            sheet.cell(row = i+2, column = 2).value = accessories[i].manufacturer.name
        sheet.cell(row = i+2, column = 3).value = accessories[i].name
        sheet.cell(row = i+2, column = 4).value = accessories[i].price
        if accessories[i].department:
            sheet.cell(row = i+2, column = 5).value = accessories[i].department.name
        if accessories[i].assigned_to:
            sheet.cell(row = i+2, column = 6).value = accessories[i].assigned_to.name
        if accessories[i].location:
            sheet.cell(row = i+2, column = 7).value = accessories[i].location.name
        sheet.cell(row = i+2, column = 8).value = accessories[i].purchase_date
        if accessories[i].status:
            sheet.cell(row = i+2, column = 9).value = accessories[i].status.name
        if accessories[i].supplier:
            sheet.cell(row = i+2, column = 10).value = accessories[i].supplier.name
        sheet.cell(row = i+2, column = 11).value = accessories[i].serial
        sheet.cell(row = i+2, column = 12).value = accessories[i].invoice

    workbook.save(object)

    # create a response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    # tell the browser what the file is named
    response['Content-Disposition'] = 'attachment;filename="Exported_accessories.xlsx"'
    # put the spreadsheet data into the response
    response.write(object.getvalue())
    # return the response
    return response

def export_employees(request):
    employees = Employee.objects.all()
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Employee"
    
    for i in range(len(employees)):
        sheet.cell(row = i+2, column = 1).value = employees[i].name

    workbook.save(object)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename="Exported_employees.xlsx"'
    response.write(object.getvalue())
    return response

def export_categories(request):
    categories = Category.objects.all()
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Category"
    
    for i in range(len(categories)):
        sheet.cell(row = i+2, column = 1).value = categories[i].name

    workbook.save(object)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename="Exported_categories.xlsx"'
    response.write(object.getvalue())
    return response

def export_manufacturers(request):
    manufacturers = Manufacturer.objects.all()
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Manufacturer"
    
    for i in range(len(manufacturers)):
        sheet.cell(row = i+2, column = 1).value = manufacturers[i].name

    workbook.save(object)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename="Exported_manufacturers.xlsx"'
    response.write(object.getvalue())
    return response

def export_suppliers(request):
    suppliers = Supplier.objects.all()
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Supplier"
    
    for i in range(len(suppliers)):
        sheet.cell(row = i+2, column = 1).value = suppliers[i].name

    workbook.save(object)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename="Exported_suppliers.xlsx"'
    response.write(object.getvalue())
    return response

def export_departments(request):
    departments = Department.objects.all()
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Department"
    
    for i in range(len(departments)):
        sheet.cell(row = i+2, column = 1).value = departments[i].name

    workbook.save(object)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename="Exported_departments.xlsx"'
    response.write(object.getvalue())
    return response

def export_licenses(request):
    licenses = License.objects.all()
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet["A1"] = "Software name"
    sheet["B1"] = "Licensed to"
    sheet["C1"] = "License email"
    sheet["D1"] = "License type"
    sheet["E1"] = "Seats"
    sheet["F1"] = "Reference no."
    sheet["G1"] = "Purchase date"
    sheet["H1"] = "Expiration date"
    sheet["I1"] = "Cost"
    sheet["J1"] = "Billing terms"
    sheet["K1"] = "Notes"
    
    for i in range(len(licenses)):
        sheet.cell(row = i+2, column = 1).value = licenses[i].software_name
        sheet.cell(row = i+2, column = 2).value = licenses[i].to_name
        sheet.cell(row = i+2, column = 3).value = licenses[i].to_email
        sheet.cell(row = i+2, column = 4).value = licenses[i].get_license_type_display()
        sheet.cell(row = i+2, column = 5).value = licenses[i].seats
        sheet.cell(row = i+2, column = 6).value = licenses[i].reference_no
        sheet.cell(row = i+2, column = 7).value = licenses[i].purchase_date
        sheet.cell(row = i+2, column = 8).value = licenses[i].expiration_date
        sheet.cell(row = i+2, column = 9).value = licenses[i].cost
        sheet.cell(row = i+2, column = 10).value = licenses[i].billing_terms
        sheet.cell(row = i+2, column = 11).value = licenses[i].notes

    workbook.save(object)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename="Exported_licenses.xlsx"'
    response.write(object.getvalue())
    return response

def export_activity(request):
    activity = Activity.objects.all().order_by('-id')
    object = BytesIO()

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet["A1"] = "ID"
    sheet["B1"] = "Timestamp"
    sheet["C1"] = "Type"
    sheet["D1"] = "Event"
    sheet["E1"] = "Asset"
    sheet["F1"] = "Employee"
    sheet["G1"] = "Notes"
    sheet["H1"] = "User"
    
    for i in range(len(activity)):
        sheet.cell(row = i+2, column = 1).value = str(activity[i].id)
        sheet.cell(row = i+2, column = 2).value = activity[i].timestamp.replace(tzinfo=None)
        sheet.cell(row = i+2, column = 3).value = activity[i].get_type_display()
        sheet.cell(row = i+2, column = 4).value = activity[i].event
        sheet.cell(row = i+2, column = 5).value = activity[i].asset_string
        sheet.cell(row = i+2, column = 6).value = str(activity[i].employee)
        sheet.cell(row = i+2, column = 7).value = activity[i].notes
        if activity[i].user:
            sheet.cell(row = i+2, column = 8).value = activity[i].user.get_full_name()

    workbook.save(object)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename="Exported_activity.xlsx"'
    response.write(object.getvalue())
    return response